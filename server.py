import os, json, urllib.parse, urllib.request, hashlib, io
from http.server import HTTPServer, BaseHTTPRequestHandler
import psycopg2, psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PORT = int(os.environ.get('PORT', 8082))
DATABASE_URL = os.environ.get('DATABASE_URL', '')
GOOGLE_CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID', '')
BREVO_API_KEY = os.environ.get('BREVO_API_KEY', '')
MSG91_AUTH_KEY = os.environ.get('MSG91_AUTH_KEY', '')
MSG91_SENDER = os.environ.get('MSG91_SENDER', '')
ADMIN_EMAILS = [e.strip().lower() for e in os.environ.get('ADMIN_EMAILS', 'hr@cuemath.com,ajay.yadav@cuemath.com,ravindra.rajput@cuemath.com,deepak.tyagi@cuemath.com').split(',')]
ALLOWED_DOMAIN = 'cuemath.com'


def get_db():
    if not DATABASE_URL:
        raise Exception('DATABASE_URL not set')
    return psycopg2.connect(DATABASE_URL, connect_timeout=30)


def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS routes (
            id SERIAL PRIMARY KEY,
            route_name TEXT NOT NULL,
            description TEXT DEFAULT '',
            pick_location TEXT DEFAULT '',
            drop_location TEXT DEFAULT '',
            pick_time TEXT DEFAULT '',
            drop_time TEXT DEFAULT '',
            report_time TEXT DEFAULT '',
            cab_cost NUMERIC DEFAULT 0,
            guard_cost NUMERIC DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS vendors (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            type TEXT NOT NULL,
            email TEXT DEFAULT '',
            mobile TEXT DEFAULT '',
            password_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS night_employees (
            id SERIAL PRIMARY KEY,
            emp_name TEXT NOT NULL,
            emp_code TEXT DEFAULT '',
            email TEXT NOT NULL,
            mobile TEXT DEFAULT '',
            department TEXT DEFAULT '',
            pickup_address TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS roster (
            id SERIAL PRIMARY KEY,
            employee_id INTEGER REFERENCES night_employees(id) ON DELETE CASCADE,
            route_id INTEGER REFERENCES routes(id) ON DELETE CASCADE,
            shift_date TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(employee_id, shift_date)
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS cab_assignments (
            id SERIAL PRIMARY KEY,
            roster_id INTEGER REFERENCES roster(id) ON DELETE CASCADE UNIQUE,
            vendor_id INTEGER REFERENCES vendors(id),
            driver_name TEXT DEFAULT '',
            driver_mobile TEXT DEFAULT '',
            vehicle_type TEXT DEFAULT '',
            vehicle_number TEXT DEFAULT '',
            assigned_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS security_assignments (
            id SERIAL PRIMARY KEY,
            roster_id INTEGER REFERENCES roster(id) ON DELETE CASCADE UNIQUE,
            vendor_id INTEGER REFERENCES vendors(id),
            guard_name TEXT DEFAULT '',
            guard_mobile TEXT DEFAULT '',
            assigned_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    # Migrations
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS guard_cost NUMERIC DEFAULT 0")
    cur.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    ''')
    cur.execute("INSERT INTO settings (key, value) VALUES ('snack_cost_per_emp', '0') ON CONFLICT (key) DO NOTHING")
    cur.execute('''
        CREATE TABLE IF NOT EXISTS inventory_items (
            id SERIAL PRIMARY KEY,
            item_name TEXT NOT NULL,
            unit TEXT DEFAULT 'pcs',
            current_stock NUMERIC DEFAULT 0,
            min_stock NUMERIC DEFAULT 5,
            cost_per_unit NUMERIC DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS inventory_log (
            id SERIAL PRIMARY KEY,
            item_id INTEGER REFERENCES inventory_items(id) ON DELETE CASCADE,
            log_date TEXT NOT NULL,
            quantity NUMERIC NOT NULL,
            log_type TEXT DEFAULT 'use',
            note TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    ''')
    cur.execute("ALTER TABLE inventory_items ADD COLUMN IF NOT EXISTS purchase_date TEXT DEFAULT ''")
    cur.execute("ALTER TABLE inventory_log ADD COLUMN IF NOT EXISTS cost_per_unit NUMERIC DEFAULT 0")
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS default_driver_name TEXT DEFAULT ''")
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS default_driver_mobile TEXT DEFAULT ''")
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS default_vehicle_type TEXT DEFAULT ''")
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS default_vehicle_number TEXT DEFAULT ''")
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS default_guard_name TEXT DEFAULT ''")
    cur.execute("ALTER TABLE routes ADD COLUMN IF NOT EXISTS default_guard_mobile TEXT DEFAULT ''")
    cur.execute("""
        ALTER TABLE night_employees ADD COLUMN IF NOT EXISTS pickup_address TEXT DEFAULT ''
    """)
    conn.commit()
    cur.close()
    conn.close()


def verify_google_token_debug(token):
    import base64, time
    try:
        parts = token.split('.')
        if len(parts) != 3:
            return None, 'Invalid JWT structure'
        padding = 4 - len(parts[1]) % 4
        payload_bytes = base64.urlsafe_b64decode(parts[1] + '=' * padding)
        data = json.loads(payload_bytes.decode('utf-8'))
        iss = data.get('iss', '')
        if iss not in ('accounts.google.com', 'https://accounts.google.com'):
            return None, 'Bad issuer: ' + iss
        aud = data.get('aud', '')
        if aud != GOOGLE_CLIENT_ID:
            return None, 'Aud mismatch: got=' + str(aud) + ' want=' + GOOGLE_CLIENT_ID
        if data.get('exp', 0) < time.time():
            return None, 'Token expired'
        email = data.get('email', '').lower()
        domain = email.split('@')[-1] if '@' in email else ''
        name = data.get('name', email)
        if domain == ALLOWED_DOMAIN:
            role = 'admin' if email in ADMIN_EMAILS else 'employee'
            return {'email': email, 'name': name, 'role': role}, None
        return None, 'Domain rejected: ' + domain
    except Exception as ex:
        return None, str(ex)

def verify_google_token(token):
    user, _ = verify_google_token_debug(token)
    return user


def send_email(to_email, to_name, subject, html):
    if not BREVO_API_KEY or not to_email:
        return
    payload = json.dumps({
        "sender": {"name": "Cuemath Cab Service", "email": "hr@cuemath.com"},
        "to": [{"email": to_email, "name": to_name}],
        "subject": subject,
        "htmlContent": html
    }).encode()
    req = urllib.request.Request(
        'https://api.brevo.com/v3/smtp/email',
        data=payload,
        headers={'api-key': BREVO_API_KEY, 'Content-Type': 'application/json'}
    )
    try:
        urllib.request.urlopen(req)
    except Exception as e:
        print('Email error:', e)


def send_whatsapp(mobile, message):
    if not MSG91_AUTH_KEY or not mobile:
        return
    mobile = ''.join(filter(str.isdigit, mobile))
    if len(mobile) == 10:
        mobile = '91' + mobile
    payload = json.dumps({
        "sender": MSG91_SENDER,
        "mobiles": mobile,
        "message": message
    }).encode()
    req = urllib.request.Request(
        'https://api.msg91.com/api/sendhttp.php',
        data=payload,
        headers={'authkey': MSG91_AUTH_KEY, 'Content-Type': 'application/json'}
    )
    try:
        urllib.request.urlopen(req)
    except Exception as e:
        print('WhatsApp error:', e)


def notify_assignment(roster_id, conn):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute('''
        SELECT r.shift_date,
               ne.emp_name, ne.email, ne.mobile,
               ro.route_name, ro.pick_location, ro.drop_location, ro.pick_time, ro.drop_time,
               ca.driver_name, ca.driver_mobile, ca.vehicle_type, ca.vehicle_number,
               sa.guard_name, sa.guard_mobile
        FROM roster r
        JOIN night_employees ne ON r.employee_id = ne.id
        JOIN routes ro ON r.route_id = ro.id
        LEFT JOIN cab_assignments ca ON ca.roster_id = r.id
        LEFT JOIN security_assignments sa ON sa.roster_id = r.id
        WHERE r.id = %s
    ''', (roster_id,))
    e = cur.fetchone()
    cur.close()
    if not e:
        return

    lines = [
        '<b>Cab Booking Confirmed!</b>',
        f'Date: {e["shift_date"]}',
        f'Route: {e["route_name"]}',
        f'Pick Up: {e["pick_location"]} | {e["pick_time"]}',
        f'Drop: {e["drop_location"]} | {e["drop_time"]}',
    ]
    if e['driver_name']:
        lines.append(f'Driver: {e["driver_name"]} | {e["driver_mobile"]}')
        lines.append(f'Vehicle: {e["vehicle_type"]} | {e["vehicle_number"]}')
    if e['guard_name']:
        lines.append(f'Guard: {e["guard_name"]} | {e["guard_mobile"]}')

    html_body = '<br>'.join(lines)
    html = f'<div style="font-family:sans-serif;font-size:15px;line-height:2">{html_body}</div>'
    plain = '\n'.join(l.replace('<b>', '').replace('</b>', '') for l in lines)

    send_email(e['email'], e['emp_name'], 'Cab Booking Confirmed', html)
    send_whatsapp(e['mobile'], plain)
    for admin_email in ADMIN_EMAILS:
        send_email(admin_email, 'Admin', f'Cab assigned: {e["emp_name"]} | {e["shift_date"]}', html)


def get_roster_full(cur, where_clause='', params=()):
    cur.execute(f'''
        SELECT r.id, r.shift_date, r.employee_id, r.route_id,
               ne.emp_name, ne.emp_code, ne.email, ne.mobile, ne.department,
               ro.route_name, ro.pick_location, ro.drop_location, ro.pick_time, ro.drop_time, ro.cab_cost, ro.guard_cost,
               ca.id as cab_id, ca.driver_name, ca.driver_mobile, ca.vehicle_type, ca.vehicle_number, ca.vendor_id as cab_vendor_id,
               sa.id as sec_id, sa.guard_name, sa.guard_mobile, sa.vendor_id as sec_vendor_id
        FROM roster r
        JOIN night_employees ne ON r.employee_id = ne.id
        JOIN routes ro ON r.route_id = ro.id
        LEFT JOIN cab_assignments ca ON ca.roster_id = r.id
        LEFT JOIN security_assignments sa ON sa.roster_id = r.id
        {where_clause}
        ORDER BY r.shift_date, ro.route_name, ne.emp_name
    ''', params)
    rows = cur.fetchall()
    result = []
    for row in rows:
        d = dict(row)
        for k in d:
            if d[k] is None:
                d[k] = ''
            elif hasattr(d[k], '__float__'):
                d[k] = float(d[k])
        result.append(d)
    return result


def build_excel(rows, employees=None):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Roster ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Roster'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    alt_fill    = PatternFill('solid', fgColor='D6E4F0')
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Date', 'Employee', 'Emp Code', 'Department', 'Mobile',
        'Route', 'Pick Up', 'Drop', 'Pick Time', 'Drop Time',
        'Driver', 'Driver Mobile', 'Vehicle Type', 'Vehicle No',
        'Guard', 'Guard Mobile', 'Cab Cost (₹)'
    ]
    col_widths = [12, 22, 10, 16, 14, 18, 22, 22, 10, 10, 18, 14, 14, 14, 18, 14, 12]

    ws.append(headers)
    for i, cell in enumerate(ws[1], 1):
        cell.font = header_fill and header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1]
    ws.row_dimensions[1].height = 30

    for idx, r in enumerate(rows, 2):
        row = [
            str(r.get('shift_date', '')),
            r.get('emp_name', ''), r.get('emp_code', ''), r.get('department', ''), r.get('mobile', ''),
            r.get('route_name', ''), r.get('pick_location', ''), r.get('drop_location', ''),
            r.get('pick_time', ''), r.get('drop_time', ''),
            r.get('driver_name', ''), r.get('driver_mobile', ''),
            r.get('vehicle_type', ''), r.get('vehicle_number', ''),
            r.get('guard_name', ''), r.get('guard_mobile', ''),
            r.get('cab_cost', 0)
        ]
        ws.append(row)
        fill = alt_fill if idx % 2 == 0 else None
        for cell in ws[idx]:
            if fill:
                cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Date-wise Cost Summary ──────────────────────────────
    ws2 = wb.create_sheet('Date-wise Cost')
    gold_fill   = PatternFill('solid', fgColor='FFF2CC')
    bold_font   = Font(bold=True, size=10)

    # Collect all unique routes (sorted)
    all_routes = sorted(set(r.get('route_name', '') for r in rows))

    # date → { route → {employees, cab_cost} }
    date_map = {}
    for r in rows:
        date  = str(r.get('shift_date', ''))
        route = r.get('route_name', 'Unknown')
        cost  = float(r.get('cab_cost', 0) or 0)
        if date not in date_map:
            date_map[date] = {}
        if route not in date_map[date]:
            date_map[date][route] = {'employees': 0, 'cost': cost}
        date_map[date][route]['employees'] += 1

    # Header row: Date | Route1 Emp | Route1 Cost | Route2 Emp | ... | Total Emp | Total Cost
    hdr = ['Date']
    for rt in all_routes:
        hdr += [f'{rt}\nEmployees', f'{rt}\nCost (₹)']
    hdr += ['Total Employees', 'Total Cost (₹)']
    ws2.append(hdr)

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws2.row_dimensions[1].height = 35
    ws2.column_dimensions['A'].width = 13
    for ci in range(2, len(hdr) + 1):
        ws2.column_dimensions[get_column_letter(ci)].width = 14

    grand_emp = 0
    grand_cost = 0
    for idx, date in enumerate(sorted(date_map.keys()), 2):
        row = [date]
        day_emp = 0
        day_cost = 0
        for rt in all_routes:
            d = date_map[date].get(rt, {'employees': 0, 'cost': 0})
            row += [d['employees'] or '', d['cost'] if d['employees'] else '']
            day_emp  += d['employees']
            day_cost += d['cost']
        row += [day_emp, day_cost]
        grand_emp  += day_emp
        grand_cost += day_cost
        ws2.append(row)
        fill = alt_fill if idx % 2 == 0 else None
        for cell in ws2[idx]:
            if fill: cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Grand Total row
    tr = ws2.max_row + 1
    ws2.cell(tr, 1, 'GRAND TOTAL').font = bold_font
    ws2.cell(tr, 1).fill = gold_fill
    ws2.cell(tr, len(hdr) - 1, grand_emp).font = bold_font
    ws2.cell(tr, len(hdr) - 1).fill = gold_fill
    ws2.cell(tr, len(hdr), grand_cost).font = bold_font
    ws2.cell(tr, len(hdr)).fill = gold_fill
    ws2.freeze_panes = 'B2'

    # ── Sheet 3: Employee Master ──────────────────────────────────────
    if employees:
        ws3 = wb.create_sheet('Employee Master')
        emp_headers = ['Emp Code', 'Name', 'Department', 'Mobile', 'Email', 'Pickup Address']
        emp_widths  = [12, 24, 18, 15, 30, 30]
        ws3.append(emp_headers)
        for i, cell in enumerate(ws3[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws3.column_dimensions[get_column_letter(i)].width = emp_widths[i-1]
        ws3.row_dimensions[1].height = 25
        for idx, e in enumerate(sorted(employees, key=lambda x: x.get('emp_name', '')), 2):
            ws3.append([
                e.get('emp_code', ''), e.get('emp_name', ''),
                e.get('department', ''), e.get('mobile', ''), e.get('email', ''),
                e.get('pickup_address', '')
            ])
            fill = alt_fill if idx % 2 == 0 else None
            for cell in ws3[idx]:
                if fill: cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical='center')
        ws3.freeze_panes = 'A2'
        ws3.auto_filter.ref = ws3.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def send_json(self, data, status=200):
        body = json.dumps(data, default=str).encode()
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', len(body))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, html):
        body = html if isinstance(html, bytes) else html.encode()
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.end_headers()
        self.wfile.write(body)

    def read_body(self):
        length = int(self.headers.get('Content-Length', 0))
        return json.loads(self.rfile.read(length)) if length else {}

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        params = urllib.parse.parse_qs(parsed.query)

        if path in ('/', '/index.html'):
            with open(os.path.join(os.path.dirname(__file__), 'index.html'), 'rb') as f:
                self.send_html(f.read())
            return

        if path == '/api/config':
            self.send_json({'google_client_id': GOOGLE_CLIENT_ID})
            return

        try:
            conn = get_db()
        except Exception as dbe:
            print('DB connect error (GET):', dbe)
            self.send_json({'error': 'DB unavailable: ' + str(dbe)}, 503)
            return
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        try:
            if path == '/api/admin/settings':
                cur.execute('SELECT key, value FROM settings')
                self.send_json({r['key']: r['value'] for r in cur.fetchall()})

            elif path == '/api/admin/inventory':
                cur.execute('''
                    SELECT i.*,
                        COALESCE(SUM(CASE WHEN l.log_type='add'  THEN l.quantity ELSE 0 END), 0) AS total_purchased,
                        COALESCE(SUM(CASE WHEN l.log_type='use'  THEN l.quantity ELSE 0 END), 0) AS total_used
                    FROM inventory_items i
                    LEFT JOIN inventory_log l ON l.item_id = i.id
                    GROUP BY i.id ORDER BY i.item_name
                ''')
                self.send_json([dict(r) for r in cur.fetchall()])

            elif path == '/api/admin/inventory/log':
                q = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
                df = q.get('date_from',[None])[0]; dt = q.get('date_to',[None])[0]
                if df and dt:
                    cur.execute('''SELECT l.*, i.item_name, i.unit, i.cost_per_unit as item_cost_per_unit FROM inventory_log l
                        JOIN inventory_items i ON l.item_id=i.id
                        WHERE l.log_date BETWEEN %s AND %s ORDER BY l.log_date DESC''', (df, dt))
                else:
                    cur.execute('''SELECT l.*, i.item_name, i.unit, i.cost_per_unit as item_cost_per_unit FROM inventory_log l
                        JOIN inventory_items i ON l.item_id=i.id ORDER BY l.log_date DESC LIMIT 100''')
                self.send_json([dict(r) for r in cur.fetchall()])

            elif path == '/api/admin/routes':
                cur.execute('SELECT * FROM routes ORDER BY route_name')
                self.send_json([dict(r) for r in cur.fetchall()])

            elif path == '/api/admin/employees':
                cur.execute('SELECT * FROM night_employees ORDER BY emp_name')
                self.send_json([dict(r) for r in cur.fetchall()])

            elif path == '/api/admin/vendors':
                cur.execute("SELECT id, name, type, email, mobile FROM vendors ORDER BY type, name")
                self.send_json([dict(r) for r in cur.fetchall()])

            elif path == '/api/admin/roster':
                date_from = params.get('date_from', [None])[0]
                date_to = params.get('date_to', [None])[0]
                if date_from and date_to:
                    rows = get_roster_full(cur, 'WHERE r.shift_date BETWEEN %s AND %s', (date_from, date_to))
                else:
                    rows = get_roster_full(cur)
                self.send_json(rows)

            elif path == '/api/vendor/roster':
                vendor_id = params.get('vendor_id', [None])[0]
                vendor_type = params.get('type', [None])[0]
                date_from = params.get('date_from', [None])[0]
                date_to = params.get('date_to', [None])[0]
                if date_from and date_to:
                    rows = get_roster_full(cur, 'WHERE r.shift_date BETWEEN %s AND %s', (date_from, date_to))
                else:
                    rows = get_roster_full(cur)
                self.send_json(rows)

            elif path == '/api/employee/plan':
                email = params.get('email', [None])[0]
                if email:
                    rows = get_roster_full(cur, 'WHERE LOWER(ne.email) = LOWER(%s)', (email,))
                    self.send_json(rows)
                else:
                    self.send_json([])

            elif path == '/api/admin/export':
                date_from = params.get('date_from', [None])[0]
                date_to   = params.get('date_to',   [None])[0]
                if date_from and date_to:
                    rows = get_roster_full(cur, 'WHERE r.shift_date BETWEEN %s AND %s', (date_from, date_to))
                else:
                    rows = get_roster_full(cur)
                cur.execute('SELECT * FROM night_employees ORDER BY emp_name')
                employees = [dict(r) for r in cur.fetchall()]
                xlsx = build_excel(rows, employees)
                fname = f'cab_roster_{date_from or "all"}_to_{date_to or "all"}.xlsx'
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
                self.send_header('Content-Length', len(xlsx))
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(xlsx)

            else:
                self.send_json({'error': 'Not found'}, 404)

        finally:
            cur.close()
            conn.close()

    def do_POST(self):
        path = urllib.parse.urlparse(self.path).path
        body = self.read_body()

        # Auth endpoints don't need DB
        if path == '/api/auth/google':
            token = body.get('token', '')
            print('Auth attempt, token len:', len(token))
            user, dbg = verify_google_token_debug(token)
            if user:
                self.send_json({'success': True, 'user': user})
            else:
                self.send_json({'success': False, 'error': 'Unauthorized', 'debug': dbg}, 401)
            return

        try:
            conn = get_db()
        except Exception as dbe:
            print('DB connect error (POST):', dbe)
            self.send_json({'success': False, 'error': 'DB unavailable: ' + str(dbe)}, 503)
            return
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        try:
            if path == '/api/admin/inventory/add':
                cur.execute('''INSERT INTO inventory_items (item_name, unit, current_stock, min_stock, cost_per_unit, purchase_date)
                    VALUES (%s,%s,%s,%s,%s,%s) RETURNING id''',
                    (body['item_name'], body.get('unit','pcs'), body.get('current_stock',0),
                     body.get('min_stock',5), body.get('cost_per_unit',0), body.get('purchase_date','')))
                new_id = cur.fetchone()['id']
                conn.commit()
                self.send_json({'success': True, 'id': new_id})

            elif path == '/api/admin/inventory/update':
                cur.execute('''UPDATE inventory_items SET item_name=%s, unit=%s, current_stock=%s, min_stock=%s, cost_per_unit=%s, purchase_date=%s WHERE id=%s''',
                    (body['item_name'], body.get('unit','pcs'), body.get('current_stock',0),
                     body.get('min_stock',5), body.get('cost_per_unit',0), body.get('purchase_date',''), body['id']))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/inventory/delete':
                cur.execute('DELETE FROM inventory_items WHERE id=%s', (body['id'],))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/inventory/log/delete':
                log_id = body['id']
                cur.execute('SELECT item_id, quantity, log_type FROM inventory_log WHERE id=%s', (log_id,))
                log = cur.fetchone()
                if log:
                    if log['log_type'] == 'add':
                        cur.execute('UPDATE inventory_items SET current_stock=GREATEST(0,current_stock-%s) WHERE id=%s', (log['quantity'], log['item_id']))
                    else:
                        cur.execute('UPDATE inventory_items SET current_stock=current_stock+%s WHERE id=%s', (log['quantity'], log['item_id']))
                    cur.execute('DELETE FROM inventory_log WHERE id=%s', (log_id,))
                    conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/inventory/log':
                # Add stock in or use out
                item_id = body['item_id']
                qty = float(body['quantity'])
                log_type = body.get('log_type', 'use')
                if log_type == 'add':
                    cur.execute('''UPDATE inventory_items SET current_stock=current_stock+%s,
                        purchase_date=CASE WHEN purchase_date IS NULL OR purchase_date='' THEN %s ELSE purchase_date END
                        WHERE id=%s''', (qty, body.get('log_date', ''), item_id))
                else:
                    cur.execute('UPDATE inventory_items SET current_stock=GREATEST(0,current_stock-%s) WHERE id=%s', (qty, item_id))
                cost = float(body.get('cost_per_unit', 0))
                cur.execute('''INSERT INTO inventory_log (item_id, log_date, quantity, log_type, note, cost_per_unit)
                    VALUES (%s,%s,%s,%s,%s,%s)''',
                    (item_id, body.get('log_date', ''), qty, log_type, body.get('note',''), cost))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/settings':
                for key, value in body.items():
                    cur.execute("INSERT INTO settings (key,value) VALUES (%s,%s) ON CONFLICT (key) DO UPDATE SET value=%s", (key, str(value), str(value)))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/auth/vendor':
                name = body.get('name', '').strip()
                pw = body.get('password', '')
                ph = hash_pw(pw)
                cur.execute('SELECT id, name, type FROM vendors WHERE LOWER(name)=LOWER(%s) AND password_hash=%s', (name, ph))
                vendor = cur.fetchone()
                if vendor:
                    self.send_json({'success': True, 'vendor': dict(vendor)})
                else:
                    self.send_json({'success': False, 'error': 'Invalid credentials'}, 401)

            elif path == '/api/admin/routes':
                cur.execute('''
                    INSERT INTO routes (route_name, description, pick_location, drop_location, pick_time, drop_time, report_time, cab_cost, guard_cost, default_driver_name, default_driver_mobile, default_vehicle_type, default_vehicle_number, default_guard_name, default_guard_mobile)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                ''', (body['route_name'], body.get('description',''), body.get('pick_location',''),
                      body.get('drop_location',''), body.get('pick_time',''), body.get('drop_time',''),
                      body.get('report_time',''), body.get('cab_cost', 0), body.get('guard_cost', 0),
                      body.get('default_driver_name',''), body.get('default_driver_mobile',''),
                      body.get('default_vehicle_type',''), body.get('default_vehicle_number',''),
                      body.get('default_guard_name',''), body.get('default_guard_mobile','')))
                new_id = cur.fetchone()['id']
                conn.commit()
                self.send_json({'success': True, 'id': new_id})

            elif path == '/api/admin/routes/update':
                cur.execute('''
                    UPDATE routes SET route_name=%s, description=%s, pick_location=%s, drop_location=%s,
                    pick_time=%s, drop_time=%s, report_time=%s, cab_cost=%s, guard_cost=%s,
                    default_driver_name=%s, default_driver_mobile=%s, default_vehicle_type=%s,
                    default_vehicle_number=%s, default_guard_name=%s, default_guard_mobile=%s WHERE id=%s
                ''', (body['route_name'], body.get('description',''), body.get('pick_location',''),
                      body.get('drop_location',''), body.get('pick_time',''), body.get('drop_time',''),
                      body.get('report_time',''), body.get('cab_cost', 0), body.get('guard_cost', 0),
                      body.get('default_driver_name',''), body.get('default_driver_mobile',''),
                      body.get('default_vehicle_type',''), body.get('default_vehicle_number',''),
                      body.get('default_guard_name',''), body.get('default_guard_mobile',''), body['id']))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/routes/delete':
                cur.execute('DELETE FROM routes WHERE id=%s', (body['id'],))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/employees':
                cur.execute('''
                    INSERT INTO night_employees (emp_name, emp_code, email, mobile, department, pickup_address)
                    VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
                ''', (body['emp_name'], body.get('emp_code',''), body['email'],
                      body.get('mobile',''), body.get('department',''), body.get('pickup_address','')))
                new_id = cur.fetchone()['id']
                conn.commit()
                self.send_json({'success': True, 'id': new_id})

            elif path == '/api/admin/employees/update':
                cur.execute('''
                    UPDATE night_employees SET emp_name=%s, emp_code=%s, email=%s, mobile=%s, department=%s, pickup_address=%s
                    WHERE id=%s
                ''', (body['emp_name'], body.get('emp_code',''), body['email'],
                      body.get('mobile',''), body.get('department',''), body.get('pickup_address',''), body['id']))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/employees/delete':
                cur.execute('DELETE FROM night_employees WHERE id=%s', (body['id'],))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/vendors':
                ph = hash_pw(body['password'])
                cur.execute('''
                    INSERT INTO vendors (name, type, email, mobile, password_hash)
                    VALUES (%s,%s,%s,%s,%s) RETURNING id
                ''', (body['name'], body['type'], body.get('email',''), body.get('mobile',''), ph))
                new_id = cur.fetchone()['id']
                conn.commit()
                self.send_json({'success': True, 'id': new_id})

            elif path == '/api/admin/vendors/delete':
                cur.execute('DELETE FROM vendors WHERE id=%s', (body['id'],))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/admin/roster':
                # Bulk add roster entries for a week
                entries = body.get('entries', [])
                added = 0
                for entry in entries:
                    try:
                        cur.execute('''
                            INSERT INTO roster (employee_id, route_id, shift_date)
                            VALUES (%s,%s,%s) ON CONFLICT (employee_id, shift_date) DO NOTHING
                            RETURNING id
                        ''', (entry['employee_id'], entry['route_id'], entry['shift_date']))
                        row = cur.fetchone()
                        if row:
                            new_roster_id = row['id']
                            added += 1
                            # Auto-copy existing cab assignment for same route+date
                            cur.execute('''
                                SELECT ca.* FROM cab_assignments ca
                                JOIN roster r ON ca.roster_id = r.id
                                WHERE r.route_id=%s AND r.shift_date=%s
                                LIMIT 1
                            ''', (entry['route_id'], entry['shift_date']))
                            existing_cab = cur.fetchone()
                            if existing_cab:
                                cur.execute('''
                                    INSERT INTO cab_assignments (roster_id, vendor_id, driver_name, driver_mobile, vehicle_type, vehicle_number)
                                    VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (roster_id) DO NOTHING
                                ''', (new_roster_id, existing_cab['vendor_id'], existing_cab['driver_name'],
                                      existing_cab['driver_mobile'], existing_cab['vehicle_type'], existing_cab['vehicle_number']))
                            # Auto-copy existing guard assignment for same route+date
                            cur.execute('''
                                SELECT sa.* FROM security_assignments sa
                                JOIN roster r ON sa.roster_id = r.id
                                WHERE r.route_id=%s AND r.shift_date=%s
                                LIMIT 1
                            ''', (entry['route_id'], entry['shift_date']))
                            existing_guard = cur.fetchone()
                            if existing_guard:
                                cur.execute('''
                                    INSERT INTO security_assignments (roster_id, vendor_id, guard_name, guard_mobile)
                                    VALUES (%s,%s,%s,%s) ON CONFLICT (roster_id) DO NOTHING
                                ''', (new_roster_id, existing_guard['vendor_id'], existing_guard['guard_name'], existing_guard['guard_mobile']))
                    except Exception:
                        pass
                conn.commit()
                self.send_json({'success': True, 'added': added})

            elif path == '/api/admin/bulk-assign':
                # assignments: [{date, route_id, driver_name, driver_mobile, vehicle_type, vehicle_number, guard_name, guard_mobile}]
                assignments = body.get('assignments', [])
                cab_done = 0; guard_done = 0
                for a in assignments:
                    cur.execute('SELECT id FROM roster WHERE route_id=%s AND shift_date=%s', (a['route_id'], a['date']))
                    roster_ids = [r['id'] for r in cur.fetchall()]
                    for rid in roster_ids:
                        if a.get('driver_name','').strip():
                            cur.execute('''
                                INSERT INTO cab_assignments (roster_id, vendor_id, driver_name, driver_mobile, vehicle_type, vehicle_number)
                                VALUES (%s,NULL,%s,%s,%s,%s)
                                ON CONFLICT (roster_id) DO UPDATE SET
                                    driver_name=%s, driver_mobile=%s, vehicle_type=%s, vehicle_number=%s, assigned_at=NOW()
                            ''', (rid, a.get('driver_name',''), a.get('driver_mobile',''), a.get('vehicle_type',''), a.get('vehicle_number',''),
                                  a.get('driver_name',''), a.get('driver_mobile',''), a.get('vehicle_type',''), a.get('vehicle_number','')))
                            cab_done += 1
                        if a.get('guard_name','').strip():
                            cur.execute('''
                                INSERT INTO security_assignments (roster_id, vendor_id, guard_name, guard_mobile)
                                VALUES (%s,NULL,%s,%s)
                                ON CONFLICT (roster_id) DO UPDATE SET
                                    guard_name=%s, guard_mobile=%s, assigned_at=NOW()
                            ''', (rid, a.get('guard_name',''), a.get('guard_mobile',''),
                                  a.get('guard_name',''), a.get('guard_mobile','')))
                            guard_done += 1
                conn.commit()
                self.send_json({'success': True, 'cab_assigned': cab_done, 'guard_assigned': guard_done})

            elif path == '/api/admin/auto-assign':
                date_from = body.get('date_from')
                date_to   = body.get('date_to')
                cur.execute('''
                    SELECT r.id, ro.default_driver_name, ro.default_driver_mobile,
                           ro.default_vehicle_type, ro.default_vehicle_number,
                           ro.default_guard_name, ro.default_guard_mobile
                    FROM roster r
                    JOIN routes ro ON r.route_id = ro.id
                    WHERE r.shift_date BETWEEN %s AND %s
                ''', (date_from, date_to))
                entries = cur.fetchall()
                cab_done = 0; guard_done = 0
                for e in entries:
                    if e['default_driver_name']:
                        cur.execute('''
                            INSERT INTO cab_assignments (roster_id, vendor_id, driver_name, driver_mobile, vehicle_type, vehicle_number)
                            VALUES (%s, NULL, %s, %s, %s, %s)
                            ON CONFLICT (roster_id) DO NOTHING
                        ''', (e['id'], e['default_driver_name'], e['default_driver_mobile'],
                              e['default_vehicle_type'], e['default_vehicle_number']))
                        cab_done += 1
                    if e['default_guard_name']:
                        cur.execute('''
                            INSERT INTO security_assignments (roster_id, vendor_id, guard_name, guard_mobile)
                            VALUES (%s, NULL, %s, %s)
                            ON CONFLICT (roster_id) DO NOTHING
                        ''', (e['id'], e['default_guard_name'], e['default_guard_mobile']))
                        guard_done += 1
                conn.commit()
                self.send_json({'success': True, 'cab_assigned': cab_done, 'guard_assigned': guard_done})

            elif path == '/api/admin/roster/delete':
                cur.execute('DELETE FROM roster WHERE id=%s', (body['id'],))
                conn.commit()
                self.send_json({'success': True})

            elif path == '/api/cab/assign':
                vendor_id = body.get('vendor_id')
                roster_id = body.get('roster_id')
                # Find route_id + shift_date for this roster entry
                cur.execute('SELECT route_id, shift_date FROM roster WHERE id=%s', (roster_id,))
                rrow = cur.fetchone()
                if rrow:
                    # Get all roster IDs on same route + date
                    cur.execute('SELECT id FROM roster WHERE route_id=%s AND shift_date=%s', (rrow['route_id'], rrow['shift_date']))
                    all_ids = [r['id'] for r in cur.fetchall()]
                else:
                    all_ids = [roster_id]
                for rid in all_ids:
                    cur.execute('''
                        INSERT INTO cab_assignments (roster_id, vendor_id, driver_name, driver_mobile, vehicle_type, vehicle_number)
                        VALUES (%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (roster_id) DO UPDATE SET
                            vendor_id=%s, driver_name=%s, driver_mobile=%s,
                            vehicle_type=%s, vehicle_number=%s, assigned_at=NOW()
                    ''', (
                        rid, vendor_id,
                        body.get('driver_name',''), body.get('driver_mobile',''),
                        body.get('vehicle_type',''), body.get('vehicle_number',''),
                        vendor_id,
                        body.get('driver_name',''), body.get('driver_mobile',''),
                        body.get('vehicle_type',''), body.get('vehicle_number','')
                    ))
                conn.commit()
                notify_assignment(roster_id, conn)
                self.send_json({'success': True, 'applied_to': len(all_ids)})

            elif path == '/api/security/assign':
                vendor_id = body.get('vendor_id')
                roster_id = body.get('roster_id')
                # Find route_id + shift_date for this roster entry
                cur.execute('SELECT route_id, shift_date FROM roster WHERE id=%s', (roster_id,))
                rrow = cur.fetchone()
                if rrow:
                    cur.execute('SELECT id FROM roster WHERE route_id=%s AND shift_date=%s', (rrow['route_id'], rrow['shift_date']))
                    all_ids = [r['id'] for r in cur.fetchall()]
                else:
                    all_ids = [roster_id]
                for rid in all_ids:
                    cur.execute('''
                        INSERT INTO security_assignments (roster_id, vendor_id, guard_name, guard_mobile)
                        VALUES (%s,%s,%s,%s)
                        ON CONFLICT (roster_id) DO UPDATE SET
                            vendor_id=%s, guard_name=%s, guard_mobile=%s, assigned_at=NOW()
                    ''', (
                        rid, vendor_id,
                        body.get('guard_name',''), body.get('guard_mobile',''),
                        vendor_id,
                        body.get('guard_name',''), body.get('guard_mobile','')
                    ))
                conn.commit()
                notify_assignment(roster_id, conn)
                self.send_json({'success': True, 'applied_to': len(all_ids)})

            else:
                self.send_json({'error': 'Not found'}, 404)

        except Exception as err:
            conn.rollback()
            self.send_json({'success': False, 'error': str(err)}, 500)
        finally:
            cur.close()
            conn.close()


if __name__ == '__main__':
    try:
        init_db()
        print('Database initialized')
    except Exception as e:
        print(f'DB init warning: {e}')
    print(f'Server running on port {PORT}')
    HTTPServer(('0.0.0.0', PORT), Handler).serve_forever()
